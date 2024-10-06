import openpyxl


def save_to_excel(data):
    try:
        wb = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if 'Список товаров' not in wb.sheetnames:
        ws = wb.create_sheet('Список товаров')
        ws.append(['Название', 'Цена со скидкой', 'Цена без скидки', 'Количество', 'Ссылка', 'Ссылка на профиль', 'Имя', 'Дата'])
    else:
        ws = wb['Список товаров']

    ws.append(data)
    wb.save('data.xlsx')


def delete_inf():
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb['Список товаров']
    wb.remove_sheet(sheet)
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         cell.value = None
    #         # cell.delete_blank_columns()
    # sheet.cells.delete_blank_columns()
    ws = wb.create_sheet('Список товаров')
    ws.append(['Название', 'Цена со скидкой', 'Цена без скидки', 'Количество', 'Ссылка', 'Ссылка на профиль', 'Имя', 'Дата'])

        # Сохранение изменений в книге
    wb.save('data.xlsx')
    return f"Лист {sheet} был успешно очищен."


def delete_last_inf(user_link):
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb['Список товаров']
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == user_link:
                r, c = cell.row, cell.column
                sheet.delete_rows(r)
    wb.save('data.xlsx')


def show_all_req(user_link):
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb['Список товаров']
    all = list()
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell == user_link:
                # print(row)
                all.append(f'Cсылка: {row[3]}, \n кол-во: {row[2]}, \n название: {row[0]}')

    text ='Ваши запросы: \n'
    for el in all:
        text += f'1) {el}\n'
    return text

